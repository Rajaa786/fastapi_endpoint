import shutil
import os
import pandas as pd
from openpyxl.styles import Font
import logging
from openpyxl import Workbook, load_workbook
import sys

bold_font = Font(bold=True)
pd.options.display.float_format = "{:,.2f}".format
pd.set_option("display.max_columns", None)
pd.set_option("display.max_rows", None)
pd.set_option("display.width", None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "../../..")))

from common_functions import CommonFunctions


class CABankStatement:
    def __init__(
        self,
        bank_names,
        pdf_paths,
        pdf_passwords,
        start_date,
        end_date,
        CA_ID,
        progress_data,
    ):
        self.writer = None
        self.bank_names = bank_names
        self.pdf_paths = pdf_paths
        self.pdf_passwords = pdf_passwords
        self.start_date = start_date
        self.end_date = end_date
        self.account_number = ""
        self.file_name = None
        self.CA_ID = CA_ID
        self.commoner = CommonFunctions(
            bank_names, pdf_paths, pdf_passwords, start_date, end_date, CA_ID
        )
        self.progress_function = progress_data.get("progress_func")
        self.current_progress = progress_data.get("current_progress")
        self.total_progress = progress_data.get("total_progress")

    def CA_Bank_statement(self, filename, dfs, name_dfs):
        data = []
        # num_pairs = len(pd.Series(dfs).to_dict())

        for key, value in name_dfs.items():
            bank_name = key
            acc_name = value[0]
            acc_num = value[1]
            if str(acc_num) == "None":
                masked_acc_num = "None"
            else:
                masked_acc_num = "X" * (len(acc_num) - 4) + acc_num[-4:]
            data.append([masked_acc_num, acc_name, bank_name])
            for item in data:
                item[2] = "".join(
                    character for character in item[2] if character.isalpha()
                )

        name_n_num_df = self.commoner.process_name_n_num_df(data)

        initial_df = pd.concat(list(dfs.values())).fillna("").reset_index(drop=True)
        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Categorizing transactions",
        )
        self.current_progress += 1
        df = self.commoner.category_add_ca(initial_df)
        new_tran_df = self.commoner.another_method(df)

        eod_sheet_df = self.commoner.eod(df)

        opening_bal, closing_bal = self.commoner.opening_and_closing_bal(
            eod_sheet_df, df
        )
        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating summary sheet"
        )
        self.current_progress += 1
        sheet_name = "Summary"
        summary_df_list = self.commoner.summary_sheet(
            df, opening_bal, closing_bal, new_tran_df
        )

        name_n_num_df.to_excel(self.writer, sheet_name=sheet_name, index=False)
        summary_df_list[0].to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=name_n_num_df.shape[0] + 2,
            index=False,
        )
        summary_df_list[1].to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=name_n_num_df.shape[0] + summary_df_list[0].shape[0] + 4,
            index=False,
        )
        summary_df_list[2].to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=name_n_num_df.shape[0]
            + summary_df_list[0].shape[0]
            + summary_df_list[1].shape[0]
            + 6,
            index=False,
        )
        summary_df_list[3].to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=name_n_num_df.shape[0]
            + summary_df_list[0].shape[0]
            + summary_df_list[1].shape[0]
            + summary_df_list[2].shape[0]
            + 8,
            index=False,
        )

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Transaction Sheet",
        )
        self.current_progress += 1
        transaction_sheet_df = self.commoner.transaction_sheet(df)
        transaction_sheet_df["Value Date"] = pd.to_datetime(
            transaction_sheet_df["Value Date"]
        ).dt.strftime("%d-%m-%Y")

        new_tran_df = self.commoner.another_method(transaction_sheet_df)
        new_tran_df.to_excel(self.writer, sheet_name="Transactions", index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating EOD Balance Sheet",
        )
        self.current_progress += 1
        eod_sheet_df.to_excel(self.writer, sheet_name="EOD Balance", index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Investment Sheet",
        )
        self.current_progress += 1
        investment_df = self.commoner.total_investment(new_tran_df)
        investment_df.to_excel(self.writer, sheet_name="Investment", index=False)

        # redemption_investment_df = self.commoner.redemption_investment(transaction_sheet_df)
        # redemption_investment_df.to_excel(self.writer, sheet_name='Redemption of Investment', index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Creditors Sheet",
        )
        self.current_progress += 1
        creditor_df = self.commoner.creditor_list(transaction_sheet_df)
        creditor_df.to_excel(self.writer, sheet_name="Creditors", index=False)

        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating Debtors Sheet"
        )
        self.current_progress += 1
        debtor_df = self.commoner.debtor_list(transaction_sheet_df)
        debtor_df.to_excel(self.writer, sheet_name="Debtors", index=False)

        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating UPI-CR Sheet"
        )
        self.current_progress += 1
        upi_cr_df = transaction_sheet_df[
            (transaction_sheet_df["Description"].str.contains("UPI", case=False))
            & (transaction_sheet_df["Credit"] > 0)
        ]
        upi_cr_df.to_excel(self.writer, sheet_name="UPI-CR", index=False)

        # Create UPI-DR sheet
        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating UPI-DR Sheet"
        )
        self.current_progress += 1
        upi_dr_df = transaction_sheet_df[
            (transaction_sheet_df["Description"].str.contains("UPI", case=False))
            & (transaction_sheet_df["Debit"] > 0)
        ]
        upi_dr_df.to_excel(self.writer, sheet_name="UPI-DR", index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Cash Withdrawal Sheet",
        )
        self.current_progress += 1
        cash_withdrawal_df = self.commoner.cash_withdraw(new_tran_df)
        cash_withdrawal_df.to_excel(
            self.writer, sheet_name="Cash Withdrawal", index=False
        )

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Cash Deposit Sheet",
        )
        self.current_progress += 1
        cash_deposit_df = self.commoner.cash_depo(new_tran_df)
        cash_deposit_df.to_excel(self.writer, sheet_name="Cash Deposit", index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Redemption, Dividend & Interest Sheet",
        )
        self.current_progress += 1
        dividend_int_df = self.commoner.div_int(new_tran_df)
        dividend_int_df.to_excel(
            self.writer, sheet_name="Redemption, Dividend & Interest", index=False
        )

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Probable EMI Sheet",
        )
        self.current_progress += 1
        emi_df = self.commoner.emi(new_tran_df)
        emi_df.to_excel(self.writer, sheet_name="Probable EMI", index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Refund-Reversal Sheet",
        )
        self.current_progress += 1
        refund = self.commoner.refund_reversal(new_tran_df)
        refund.to_excel(self.writer, sheet_name="Refund-Reversal", index=False)

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Suspense Credit Sheet",
        )
        self.current_progress += 1
        suspense_credit_df = self.commoner.suspense_credit(new_tran_df)
        suspense_credit_df.to_excel(
            self.writer, sheet_name="Suspense Credit", index=False
        )

        self.progress_function(
            self.current_progress,
            self.total_progress,
            info=f"Generating Suspense Debit Sheet",
        )
        self.current_progress += 1
        suspense_debit_df = self.commoner.suspense_debit(new_tran_df)
        suspense_debit_df.to_excel(
            self.writer, sheet_name="Suspense Debit", index=False
        )

        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating Payment Sheet"
        )
        self.current_progress += 1
        Payment = self.commoner.payment(transaction_sheet_df)
        Payment.to_excel(self.writer, sheet_name="Payment Voucher", index=False)

        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating Receipt Sheet"
        )
        self.current_progress += 1
        Receipt = self.commoner.receipt(transaction_sheet_df)
        Receipt.to_excel(self.writer, sheet_name="Receipt Voucher", index=False)

        bank_avg_balance_df = self.commoner.calculate_fixed_day_average(eod_sheet_df)
        loan_value_df = self.commoner.process_avg_last_6_months(
            filename, bank_avg_balance_df, eod_sheet_df
        )

        return loan_value_df

    # @timer_decorator
    def start_extraction(self):
        CA_ID = self.CA_ID

        dfs = {}
        name_dfs = {}
        i = 0

        for bank in self.bank_names:
            bank = str(f"{bank}{i}")
            pdf_path = self.pdf_paths[i]
            pdf_password = self.pdf_passwords[i]
            start_date = self.start_date[i]
            end_date = self.end_date[i]

            self.progress_function(
                self.current_progress,
                self.total_progress,
                info=f"Extracting bank statement",
            )
            self.current_progress += 1
            dfs[bank], name_dfs[bank] = self.commoner.extraction_process(
                bank, pdf_path, pdf_password, start_date, end_date
            )
            self.progress_function(
                self.current_progress,
                self.total_progress,
                info=f"Extraction completed successfully",
            )
            self.current_progress += 1
            print(f"Extracted {bank} bank statement successfully")
            self.account_number += f"{name_dfs[bank][1][:4]}x{name_dfs[bank][1][-4:]}_"
            i += 1

        print("|------------------------------|")
        print(self.account_number)
        print("|------------------------------|")

        try:
            saved_excel_dir = os.path.abspath(
                os.path.join(BASE_DIR, "..", "..", "saved_excel")
            )
            if not os.path.isdir(saved_excel_dir):
                os.makedirs(saved_excel_dir, exist_ok=True)

            # Ensure filename is always defined
            filename = os.path.join(
                saved_excel_dir,
                f"{self.CA_ID}_Single_Extracted_statements_file_{self.account_number}.xlsx",
            )

            # Create the Excel writer object
            self.writer = pd.ExcelWriter(filename, engine="xlsxwriter")

        except OSError as e:
            logging.error(f"Failed to create directory {saved_excel_dir}: {e}")
            raise OSError(f"Could not create output directory: {e}")

        except Exception as e:
            logging.error(f"Error processing bank statements: {e}")
            raise

        self.writer = pd.ExcelWriter(filename, engine="xlsxwriter")
        self.progress_function(
            self.current_progress, self.total_progress, info=f"Generating Excel"
        )
        self.current_progress += 1
        loan_value_df = self.CA_Bank_statement(filename, dfs, name_dfs)
        self.writer.close()

        # adjust_excel_column_widths(filename)
        self.progress_function(
            self.current_progress, self.total_progress, info=f"Finalizing Excel"
        )
        self.current_progress += 1
        self.commoner.color_summary_sheet(filename)
        self.commoner.format_numbers_with_commas(filename)
        sheet_specs = {
            "Summary": {
                "A": 50,
                "B": 15,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 15,
                "G": 15,
                "H": 15,
                "I": 15,
                "J": 15,
                "K": 15,
                "L": 15,
                "M": 15,
                "N": 15,
                "O": 15,
            },
            "DateWise Avg Balance": {
                "A": 35,
                "B": 15,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 15,
                "G": 15,
                "H": 15,
                "I": 15,
                "J": 15,
                "K": 15,
                "L": 15,
                "M": 15,
                "N": 20,
                "O": 20,
                "P": 20,
            },
            "BankWise Eligibility": {
                "A": 35,
                "B": 20,
                "C": 20,
                "D": 25,
                "E": 25,
                "F": 25,
                "G": 25,
                "H": 25,
                "I": 25,
            },
            "Transaction": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "EOD Balance": {
                "A": 15,
                "B": 15,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 15,
                "G": 15,
                "H": 15,
                "I": 15,
                "J": 15,
                "K": 15,
                "L": 15,
                "M": 15,
                "N": 15,
                "O": 15,
                "P": 15,
                "Q": 15,
                "R": 15,
                "S": 15,
                "T": 15,
                "U": 15,
                "V": 15,
                "W": 15,
                "X": 15,
                "Y": 15,
                "Z": 15,
            },
            "Probable EMI": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Bounce": {"A": 10, "B": 70, "C": 15, "D": 15, "E": 15, "F": 20, "G": 10},
            "Creditor": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Debtor": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Cash Deposit": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Cash Withdrawal": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "POS_CR": {"A": 10, "B": 70, "C": 15, "D": 15, "E": 15, "F": 20, "G": 10},
            "UPI_CR": {"A": 10, "B": 70, "C": 15, "D": 15, "E": 15, "F": 20, "G": 10},
            "Investment": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Subscription_Entertainment": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Refund-Reversal": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Suspense Credit": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Suspense Debit": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Redemption, Dividend & Interest": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Bank_charges": {
                "A": 10,
                "B": 70,
                "C": 15,
                "D": 15,
                "E": 15,
                "F": 20,
                "G": 10,
            },
            "Payment Voucher": {
                "A": 15,
                "B": 15,
                "C": 10,
                "D": 20,
                "E": 15,
                "F": 15,
                "G": 85,
            },
            "Receipt Voucher": {
                "A": 15,
                "B": 15,
                "C": 20,
                "D": 10,
                "E": 15,
                "F": 85,
            },
        }
        self.commoner.adjust_column_widths_for_varied_sheets(filename, sheet_specs)

        summary_note = (
            "Disclaimer/Caveat: The entries throughout this file and tables are based on best guess basis and "
            "information filtered under expenses and income. An attempt has been made to reflect the narration as "
            "close as possible to the actuals. \r\nHowever, variations from above are possible based on customer profile "
            "and their transactions with parties.Kindly cross check with your clients for any discrepancies"
        )
        self.commoner.Summary_note(filename, empty_rows_between=2)

        Inve_note = "*This table reflects probable transactions in securities made during the year. \r\nKindly confirm the same from Annual Information Statement (AIS) reflected on the Income Tax Portal and the capital gain reports sent by the respective authorities."
        self.commoner.Investment_note(filename, empty_rows_between=2)

        ent_note = "*This table pertains to probable entertainment expenses made during the year."
        # Entertainment_note(filename, ent_note, empty_rows_between=2)

        CreditorList = "*The entries in this table likely pertain to payments from the parties during the period mentioned. \r\nIn case of payments through online portals, we have mentioned the portal names as reflected in the narration of the bank statement. \r\nWe would like to highlight that in case of contra entries, the name of the client will be reflected as a creditor."
        self.commoner.CreditorList_note(filename, empty_rows_between=2)

        Debtor_note = "*The entries in this table likely pertains to receipts from the respective parties. \r\nIn case of receipts through online portals, we have mentioned the portal names as reflected in the narration of the bank statement. \r\nWe would like to highlight that in case of contra entries, the name of the client will be reflected as a debtor."
        self.commoner.DebtorList_note(filename, empty_rows_between=2)

        CW_note = "*The above table reflects the cash withdrawals made during the year on the basis of widely used acronyms of the finance industry."
        self.commoner.CashWithdrawalt_note(filename, CW_note, empty_rows_between=2)

        CD_note = "*The above table reflects the cash deposits made during the year on the basis of widely used acronyms of the finance industry."
        self.commoner.Cash_Deposit_note(filename, CD_note, empty_rows_between=2)

        emi_note = "* Transactions in the above table are based on the widely used acronyms of the finance industry and likely reflect EMI payment. \r\nKindly confirm the same from the loan statement or the interest certificate."
        self.commoner.Emi_note(filename, emi_note, empty_rows_between=2)

        reef_note = "*This table likely pertains to refunds/reversals/cashbacks received from card payments/online transactions."
        self.commoner.Refund_note(filename, reef_note, empty_rows_between=2)

        sus_cr_note = "*This table pertains to transactions unidentified as per the current ledger bifurcation of the software. \r\nIn case of any technical errors, inconvience is highly regretted and feedback is appreciated."
        self.commoner.Suspense_Credit_note(filename, empty_rows_between=2)

        sus_dr_note = "*This table likely pertains to transactions unidentified as per the current ledger bifurcation of the software."
        self.commoner.Suspense_Debit_note(filename, empty_rows_between=2)

        self.commoner.add_filters_to_excel(filename)
        self.commoner.create_excel_sheet(filename, loan_value_df)
        self.commoner.color_excel_tabs_inplace(filename)

        def reorder_sheets(filename):
            wb = load_workbook(filename)
            desired_order_front = ["Summary", "Opportunity to Earn"]
            existing_sheets = [
                sheet for sheet in desired_order_front if sheet in wb.sheetnames
            ]
            other_sheets = [
                sheet for sheet in wb.sheetnames if sheet not in existing_sheets
            ]
            new_order = existing_sheets + other_sheets
            wb._sheets = [wb[sheet] for sheet in new_order]
            wb.save(filename)
            return "Sheets reordered successfully"

        reorder_sheets(filename)

        folder_path = "saved_pdf"
        try:
            shutil.rmtree(folder_path)
            print(f"Removed all contents in '{folder_path}'")
        except Exception as e:
            print(f"Failed to remove '{folder_path}': {e}")

        return (self.account_number, self.current_progress)


### --------------------------------------------------------------------------------------------------------- ###
# #
# bank_names = ["AXIS"]
# pdf_paths = ["axisbank (1) (1).pdf"]
# passwords = ["PRAN004086073"]
# start_date = ["01-04-2020"]
# end_date = ["10-11-2024"]
# CA_ID = "PRANAV"
# progress_data = {
#     'progress_func': lambda current, total, info: print(f"{info} ({current}/{total})"),
#     'current_progress': 10,
#     'total_progress': 100
# }
#
# converter = CABankStatement(bank_names, pdf_paths, passwords, start_date, end_date, CA_ID, progress_data)
# converter.start_extraction()


# bank_names = ["HDFC"]
# pdf_paths = ["/Users/sanchaythalnerkar/CypherSol/accountant/banks/hdfc.pdf"]
# passwords = ["", ""]
# start_date = ["26-01-2024"]
# end_date = ["26-02-2024"]
# CA_ID = "HDFC"

# progress_data = {
#     "progress_func": lambda current, total, info: print(f"{info} ({current}/{total})"),
#     "current_progress": 10,
#     "total_progress": 100,
# }

# converter = CABankStatement(
#     bank_names, pdf_paths, passwords, start_date, end_date, CA_ID, progress_data
# )
# converter.start_extraction()
